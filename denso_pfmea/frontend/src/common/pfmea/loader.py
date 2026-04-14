from __future__ import annotations

import io
import math
import posixpath
import re
import warnings
import zipfile
from collections.abc import Iterable, Mapping, MutableMapping
from typing import Any
from xml.etree import ElementTree as ET

import pandas as pd

from src.common.concurrency import parallel_map

from .constants import (
    IGNORE_TEXTBOX_KEYWORDS,
    NS_DRAWING,
    NS_DRAWING_MAIN,
    NS_MAIN,
    NS_REL_DOC,
    NS_REL_PACK,
    PFMEA_COLUMN_MAP,
    PROCESS_DETAIL_KEYWORDS,
)
from .models import PfmeaDataset, PfmeaEntry, ProcessSummary, ShapeText
from .ratings import DEFAULT_RATING_SCALES

try:  # pragma: no cover - dependency check
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as exc:  # pragma: no cover
    raise ImportError("openpyxl is required to parse PFMEA files") from exc

# Deduplicate noisy warning: openpyxl drops unsupported WMF images on load.
warnings.filterwarnings(
    "once",
    category=UserWarning,
    message="wmf image format is not supported.*",
    module="openpyxl.reader.drawings",
)


def _clean_multiline(text: str) -> str:
    """Trim leading/trailing whitespace and collapse internal spacing."""
    collapsed = re.sub(r"[ \u3000]+", " ", text.strip())
    collapsed = re.sub(r"\s+\n", "\n", collapsed)
    collapsed = re.sub(r"\n\s+", "\n", collapsed)
    return collapsed.strip()


def _squash_line(text: str) -> str:
    """Collapse CR/LF into single spaces for final single-line fields."""
    return re.sub(r"\s+", " ", text.strip())


def _split_functions(text: str) -> tuple[str, ...]:
    if not text:
        return ()
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    normalized = normalized.replace("。", "。\n")
    segments = re.split(r"[。\n]+", normalized)
    cleaned = [_squash_line(seg).strip("。") for seg in segments if _squash_line(seg)]
    return tuple(filter(None, cleaned))


def _split_requirements(text: str) -> tuple[str, ...]:
    if not text:
        return ()
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    normalized = normalized.replace("・", "\n")
    segments = re.split(r"\n+", normalized)
    cleaned = [_squash_line(seg) for seg in segments if _squash_line(seg)]
    return tuple(cleaned)


def _split_extra_section(text: str) -> tuple[str, ...]:
    if not text:
        return ()
    normalized = text.replace("\r\n", "\n").replace("\r", "\n")
    segments = re.split(r"\n+", normalized)
    cleaned = [_squash_line(seg) for seg in segments if _squash_line(seg)]
    return tuple(cleaned)


def _parse_process_text(raw_text: str, *, anchor_row: int) -> ProcessSummary | None:
    text = raw_text.strip()
    if not text:
        return None
    if any(keyword in text for keyword in IGNORE_TEXTBOX_KEYWORDS):
        return None

    segments: MutableMapping[str, str] = {}
    current_label = "header"
    tokens = re.split(r"(【[^】]+】|\[[^\]]+\])", text)
    for token in tokens:
        if not token:
            continue
        if (
            token.startswith("【")
            and token.endswith("】")
            or token.startswith("[")
            and token.endswith("]")
        ):
            current_label = token[1:-1]
            segments.setdefault(current_label, "")
        else:
            segments.setdefault(current_label, "")
            segments[current_label] += token

    header_text = _clean_multiline(segments.get("header", ""))
    process_name = ""
    if header_text:
        match = re.search(r"No\.?\s*([^\n【\[]+)", header_text)
        if match:
            process_name = _squash_line(match.group(1))
        else:
            process_name = _squash_line(header_text)

    functions = _split_functions(segments.get("工程の機能", ""))
    requirements = _split_requirements(segments.get("製造保証項目", ""))

    extras: dict[str, tuple[str, ...]] = {}
    for key, value in segments.items():
        if key in {"header", "工程の機能", "製造保証項目"}:
            continue
        extras[key] = _split_extra_section(value)

    summary = ProcessSummary(
        process_name=process_name,
        anchor_row=anchor_row,
        raw_text=_clean_multiline(text),
        functions=functions,
        requirements=requirements,
        extra_sections=extras,
    )
    return summary


def _looks_like_process_detail(text: str) -> bool:
    if not text:
        return False
    if any(keyword in text for keyword in IGNORE_TEXTBOX_KEYWORDS):
        return False
    return any(keyword in text for keyword in PROCESS_DETAIL_KEYWORDS)


def _resolve_target(base: str, target: str) -> str:
    if target.startswith("/"):
        path = target.lstrip("/")
    else:
        base_dir = posixpath.dirname(base)
        path = posixpath.normpath(posixpath.join(base_dir, target))
    if not path.startswith("xl/"):
        path = f"xl/{path}"
    return path


def _find_nearest_summary(
    summaries: Mapping[int, ProcessSummary], target_row: int, tolerance: int = 3
) -> ProcessSummary | None:
    if not summaries:
        return None
    candidate_rows = [row for row in summaries if abs(row - target_row) <= tolerance]
    if not candidate_rows:
        return None
    closest_row = min(candidate_rows, key=lambda row: abs(row - target_row))
    return summaries.get(closest_row)


def _allocate_requirements(
    functions: tuple[str, ...], requirements: tuple[str, ...]
) -> list[tuple[str, ...]]:
    if not functions:
        return [requirements] if requirements else []

    groups: list[tuple[str, ...]] = []
    req_index = 0
    total_reqs = len(requirements)
    total_funcs = len(functions)

    for idx in range(total_funcs):
        remaining_funcs = total_funcs - idx
        remaining_reqs = total_reqs - req_index
        if remaining_funcs <= 0:
            break
        if remaining_reqs <= 0:
            groups.append(())
            continue
        share = max(1, math.ceil(remaining_reqs / remaining_funcs))
        segment = tuple(requirements[req_index : req_index + share])
        if not segment:
            segment = (requirements[req_index],)
            share = 1
        groups.append(segment)
        req_index += share

    if req_index < total_reqs and groups:
        tail = list(groups[-1]) + list(requirements[req_index:])
        groups[-1] = tuple(tail)
    while len(groups) < total_funcs:
        groups.append(())
    return groups


def _safe_int(value: Any, default: int) -> int:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(
            value
        ):  # pragma: no cover - defensive
            return default
        return int(round(value))
    text = str(value).strip()
    if not text:
        return default
    try:
        return int(float(text))
    except (ValueError, TypeError):
        return default


def _read_cell_value(ws: Worksheet, row: int, column: int) -> str:
    value = ws.cell(row=row, column=column).value
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return _squash_line(str(value))


def _find_header_row(ws: Worksheet) -> int:
    for row in range(1, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=column).value
            if cell_value and "工程No" in str(cell_value):
                return row
    raise ValueError("PFMEAヘッダー行（工程No.）が見つかりませんでした。")


def _parse_pfmea_sheet(
    ws: Worksheet,
    shape_summaries: Mapping[int, ProcessSummary],
) -> tuple[pd.DataFrame, dict[str, ProcessSummary]]:
    header_row = _find_header_row(ws)
    data_start = header_row + 2  # ヘッダ2行をスキップ

    entries: list[PfmeaEntry] = []
    process_summary_index: dict[str, ProcessSummary] = {}
    context: dict[str, str] = {
        key: "" for key in PFMEA_COLUMN_MAP if key != "priority_designation"
    }
    prev_column_values: dict[int, str] = {}

    # RPN列専用の前値キャッシュ:
    # RPN列は「〃」記号をサポートし前行値を継承する（既存PFMEA形式の仕様）。
    # 他列の「↑」記号とは異なるルールのため、別管理する。
    prev_rpn_value: int = 0

    current_process_name = ""
    current_summary: ProcessSummary | None = None

    for row_idx in range(data_start, ws.max_row + 1):
        process_cell_raw = ws.cell(row=row_idx, column=2).value
        process_cell_text = (
            str(process_cell_raw) if process_cell_raw is not None else ""
        )

        shape_summary = shape_summaries.get(row_idx)
        if shape_summary:
            if shape_summary.process_name == "↑":
                shape_summary.process_name = current_process_name
            if shape_summary.process_name and not current_process_name:
                current_process_name = shape_summary.process_name
                prev_column_values = {}
                prev_rpn_value = 0
            elif (
                shape_summary.process_name
                and shape_summary.process_name != current_process_name
            ):
                current_process_name = shape_summary.process_name
                context = dict.fromkeys(context, "")
                prev_column_values = {}
                prev_rpn_value = 0
            current_summary = shape_summary
            key = current_summary.process_name or current_process_name
            if key:
                process_summary_index[key] = current_summary
            if current_process_name and current_process_name != key:
                process_summary_index[current_process_name] = current_summary

        if _looks_like_process_detail(process_cell_text):
            summary = _parse_process_text(process_cell_text, anchor_row=row_idx)
            if summary:
                if current_process_name and not summary.process_name:
                    summary.process_name = current_process_name
                current_summary = summary
                key = current_summary.process_name or current_process_name
                if key:
                    process_summary_index[key] = current_summary
                if current_process_name and current_process_name != key:
                    process_summary_index[current_process_name] = current_summary
        else:
            process_candidate = _squash_line(process_cell_text)
            if process_candidate and process_candidate != "↑":
                current_process_name = process_candidate
                context = dict.fromkeys(context, "")
                prev_column_values = {}
                prev_rpn_value = 0
                summary = _find_nearest_summary(shape_summaries, row_idx)
                if summary:
                    if not summary.process_name:
                        summary.process_name = current_process_name
                    current_summary = summary
                    key = current_summary.process_name or current_process_name
                    if key:
                        process_summary_index[key] = current_summary
                    if current_process_name and current_process_name != key:
                        process_summary_index[current_process_name] = current_summary
                elif current_process_name in process_summary_index:
                    current_summary = process_summary_index[current_process_name]

        if not current_process_name:
            continue

        row_payload: dict[str, str] = {}
        has_any_value = False

        for field, column in PFMEA_COLUMN_MAP.items():
            value = _read_cell_value(ws, row_idx, column)
            if value == "↑":
                resolved = prev_column_values.get(column, "")
            elif value:
                resolved = value
            else:
                resolved = ""

            if not resolved and field != "priority_designation":
                resolved = context.get(field, "")

            if resolved:
                has_any_value = True
                if field != "priority_designation":
                    context[field] = resolved

            row_payload[field] = resolved
            prev_column_values[column] = resolved

        if not has_any_value:
            continue
        if not row_payload["failure_mode"]:
            continue

        if current_summary and not current_summary.process_name:
            current_summary.process_name = current_process_name
            process_summary_index[current_process_name] = current_summary

        summary_for_row = current_summary or process_summary_index.get(
            current_process_name
        )
        process_detail_text = (
            _squash_line(summary_for_row.raw_text) if summary_for_row else ""
        )
        process_functions = summary_for_row.functions if summary_for_row else ()
        process_requirements = summary_for_row.requirements if summary_for_row else ()

        # RPN: "〃"の場合は前行の値を継承、それ以外は数値化
        rpn_raw = row_payload["rpn"]
        if rpn_raw == "〃":
            rpn_value = prev_rpn_value
        else:
            rpn_value = _safe_int(rpn_raw, default=0)
            prev_rpn_value = rpn_value

        entry = PfmeaEntry(
            excel_row=row_idx,
            process_name=current_process_name,
            process_detail=process_detail_text,
            process_functions=process_functions,
            process_requirements=process_requirements,
            requirement=row_payload["requirement"],
            failure_mode=row_payload["failure_mode"],
            effect=row_payload["effect"],
            severity=_safe_int(row_payload["severity"], default=5),
            priority_designation=row_payload["priority_designation"],
            cause=row_payload["cause"],
            prevention=row_payload["prevention"],
            occurrence=_safe_int(row_payload["occurrence"], default=1),
            detection_control=row_payload["detection_control"],
            detection=_safe_int(row_payload["detection"], default=1),
            rpn=rpn_value,
            recommended_action=row_payload["recommended_action"],
            process_sheet_reflection=row_payload["process_sheet_reflection"],
            responsible_owner=row_payload["responsible_owner"],
        )
        entries.append(entry)

    if entries:
        df = pd.DataFrame([entry.__dict__ for entry in entries])
    else:
        df = pd.DataFrame(columns=list(PfmeaEntry.__annotations__.keys()))
    return df, process_summary_index


def _build_shape_summaries(
    workbook_bytes: bytes,
    sheet_targets: Mapping[str, str],
) -> dict[str, dict[int, ProcessSummary]]:
    shape_texts = _collect_shape_texts(workbook_bytes, sheet_targets)
    if not shape_texts:
        return {name: {} for name in sheet_targets}

    def _build_summary(
        task: tuple[str, list[ShapeText]],
    ) -> tuple[str, dict[int, ProcessSummary]]:
        sheet_name, shapes = task
        summary_map: dict[int, ProcessSummary] = {}
        for shape in shapes:
            summary = _parse_process_text(shape.text, anchor_row=shape.row)
            if summary:
                summary_map[shape.row] = summary
        return sheet_name, summary_map

    summary_results = parallel_map(_build_summary, list(shape_texts.items()))
    shape_summaries: dict[str, dict[int, ProcessSummary]] = dict(summary_results)
    for sheet_name in sheet_targets:
        shape_summaries.setdefault(sheet_name, {})
    return shape_summaries


def _parse_target_sheets(
    workbook_bytes: bytes,
    target_sheets: Iterable[str],
    shape_summaries: Mapping[str, dict[int, ProcessSummary]],
) -> tuple[dict[str, pd.DataFrame], dict[str, dict[str, ProcessSummary]]]:
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="wmf image format is not supported so the image is being dropped",
            category=UserWarning,
        )
        wb = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    datasets: dict[str, pd.DataFrame] = {}
    process_summaries: dict[str, dict[str, ProcessSummary]] = {}
    sheet_list = list(target_sheets)

    try:
        available_sheets = set(wb.sheetnames)
        for sheet_name in sheet_list:
            if sheet_name not in available_sheets:
                datasets.setdefault(sheet_name, pd.DataFrame())
                process_summaries.setdefault(sheet_name, {})
                continue
            ws = wb[sheet_name]
            sheet_summary_map = shape_summaries.get(sheet_name, {})
            df, summary_index = _parse_pfmea_sheet(ws, sheet_summary_map)
            datasets[sheet_name] = df
            process_summaries[sheet_name] = summary_index
    finally:
        wb.close()

    for sheet_name in sheet_list:
        datasets.setdefault(sheet_name, pd.DataFrame())
        process_summaries.setdefault(sheet_name, {})

    return datasets, process_summaries


def _add_duplicate_identifiers(df: pd.DataFrame) -> pd.DataFrame:
    """重複する要求事項に一意の識別子を付与する。"""
    if df.empty or "requirement" not in df.columns:
        return df

    # 要求事項列を正規化
    requirements = df["requirement"].fillna("").astype(str).str.strip()

    # まず、各要求事項の出現回数をカウント（空欄を除く）
    from collections import Counter

    non_empty_reqs = [req for req in requirements if req]
    occurrence_counts = Counter(non_empty_reqs)

    # 重複している要求事項のみを特定（2回以上出現）
    duplicated_reqs = {req for req, count in occurrence_counts.items() if count > 1}

    # 各重複要求事項の現在の出現番号を追跡
    seen_counts: dict[str, int] = {}
    final_requirements: list[str] = []

    for req in requirements:
        # 空欄はスキップ
        if not req:
            final_requirements.append(req)
            continue

        # 重複している要求事項の場合のみ識別子を付与
        if req in duplicated_reqs:
            if req not in seen_counts:
                seen_counts[req] = 0
            seen_counts[req] += 1
            final_requirements.append(f"{req} #{seen_counts[req]}")
        else:
            # 重複していない要求事項はそのまま
            final_requirements.append(req)

    # DataFrameを更新
    df = df.copy()
    df["requirement"] = final_requirements
    return df


def load_pfmea_bundle(workbook: bytes, sheets: Iterable[str]) -> PfmeaDataset:
    workbook_bytes = bytes(workbook)
    sheet_targets = _collect_sheet_targets(workbook_bytes)
    target_sheets = list(sheets)
    shape_summaries = _build_shape_summaries(workbook_bytes, sheet_targets)
    datasets, process_summaries = _parse_target_sheets(
        workbook_bytes, target_sheets, shape_summaries
    )

    for sheet_name in target_sheets:
        target = sheet_targets.get(sheet_name)
        if target and target in datasets:
            datasets[sheet_name] = datasets.pop(target)
            process_summaries[sheet_name] = process_summaries.pop(target)
        else:
            datasets.setdefault(sheet_name, pd.DataFrame())
            process_summaries.setdefault(sheet_name, {})

    # 重複する要求事項に識別子を付与
    for sheet_name in datasets:
        datasets[sheet_name] = _add_duplicate_identifiers(datasets[sheet_name])

    return PfmeaDataset(
        by_block=datasets,
        rating_scales=DEFAULT_RATING_SCALES,
        process_summaries=process_summaries,
    )


def _collect_sheet_targets(workbook_bytes: bytes) -> dict[str, str]:
    with zipfile.ZipFile(io.BytesIO(workbook_bytes)) as archive:
        workbook_xml = archive.read("xl/workbook.xml")
        workbook_root = ET.fromstring(workbook_xml)
        workbook_rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

    rel_map: dict[str, str] = {}
    for rel in workbook_rels.findall(f"{{{NS_REL_PACK}}}Relationship"):
        rel_id = rel.attrib.get("Id")
        target = rel.attrib.get("Target")
        if rel_id and target:
            rel_map[rel_id] = target

    sheet_targets: dict[str, str] = {}
    for sheet in workbook_root.findall(f"{{{NS_MAIN}}}sheets/{{{NS_MAIN}}}sheet"):
        name = sheet.attrib.get("name")
        rel_id = sheet.attrib.get(f"{{{NS_REL_DOC}}}id")
        if not name or not rel_id:
            continue
        target = rel_map.get(rel_id)
        if not target:
            continue
        sheet_targets[name] = _resolve_target("xl/workbook.xml", target)
    return sheet_targets


def _extract_shapes_for_sheet(
    workbook_bytes: bytes,
    sheet_name: str,
    sheet_path: str,
) -> tuple[str, list[ShapeText]]:
    ns = {"xdr": NS_DRAWING, "a": NS_DRAWING_MAIN}
    shapes: list[ShapeText] = []

    with zipfile.ZipFile(io.BytesIO(workbook_bytes)) as archive:
        sheet_dir = posixpath.dirname(sheet_path)
        rels_name = posixpath.join(
            sheet_dir, "_rels", f"{posixpath.basename(sheet_path)}.rels"
        )
        if rels_name not in archive.namelist():
            return sheet_name, shapes

        rel_root = ET.fromstring(archive.read(rels_name))
        drawing_targets: list[str] = []
        for rel in rel_root.findall(f"{{{NS_REL_PACK}}}Relationship"):
            if rel.attrib.get("Type") == f"{NS_REL_DOC}/drawing":
                target = rel.attrib.get("Target")
                if target:
                    drawing_targets.append(_resolve_target(sheet_path, target))

        if not drawing_targets:
            return sheet_name, shapes

        for drawing_path in drawing_targets:
            if drawing_path not in archive.namelist():
                continue
            drawing_root = ET.fromstring(archive.read(drawing_path))
            for anchor_tag in ("twoCellAnchor", "oneCellAnchor"):
                for anchor in drawing_root.findall(f"xdr:{anchor_tag}", ns):
                    from_node = anchor.find("xdr:from", ns)
                    if from_node is None:
                        continue
                    row_node = from_node.find("xdr:row", ns)
                    col_node = from_node.find("xdr:col", ns)
                    if row_node is None or col_node is None:
                        continue
                    row = int(row_node.text or "0") + 1
                    col = int(col_node.text or "0") + 1

                    text_fragments: list[str] = []
                    for node in anchor.iter():
                        if node.tag == f"{{{NS_DRAWING_MAIN}}}br":
                            text_fragments.append("\n")
                        elif node.tag == f"{{{NS_DRAWING_MAIN}}}t":
                            text_fragments.append(node.text or "")
                    raw_text = "".join(text_fragments).strip()
                    if not raw_text:
                        continue
                    if any(keyword in raw_text for keyword in IGNORE_TEXTBOX_KEYWORDS):
                        continue
                    shapes.append(ShapeText(row=row, col=col, text=raw_text))

    return sheet_name, shapes


def _collect_shape_texts(
    workbook_bytes: bytes, sheet_targets: Mapping[str, str]
) -> dict[str, list[ShapeText]]:
    tasks = [(name, path) for name, path in sheet_targets.items() if path]
    if not tasks:
        return {}

    def _worker(task: tuple[str, str]) -> tuple[str, list[ShapeText]]:
        sheet_name, sheet_path = task
        return _extract_shapes_for_sheet(workbook_bytes, sheet_name, sheet_path)

    results = parallel_map(_worker, tasks)
    return dict(results)


__all__ = ["load_pfmea_bundle"]
