"""Excel styling utilities for consistent formatting across exports."""

from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def _estimate_column_width(header_name: str) -> float:
    """Estimate appropriate column width based on header name and content type.

    Args:
        header_name: Column header name

    Returns:
        Estimated width in Excel units
    """
    # Mapping of column names to appropriate widths
    width_map = {
        # Metadata columns (short)
        "ブロック": 15,
        "ステーション": 18,
        "対象部品": 25,
        "変更種別": 15,
        "バリエーション": 20,
        # Process columns (medium)
        "工程名": 20,
        "製造保証項目": 25,
        # Numeric columns (narrow)
        "影響度合": 10,
        "発生度合": 10,
        "検出度合": 10,
        "重要度（RPN）": 12,
        "自信度": 10,
        "AI自信度": 10,
        # Description columns (wide)
        "工程の機能": 35,
        "要求事項（良品条件）": 35,
        "工程故障モード": 30,
        "故障の影響": 30,
        "故障の原因およびメカニズム": 40,
        # Reasoning columns (wide)
        "追加理由": 35,
        "AI追加理由": 35,
        "RPN評価理由": 40,
        "AI RPN評価理由": 40,
        # Change report specific
        "品番": 20,
        "差異理由": 40,
        "形状の特長": 30,
    }

    # Return mapped width or default
    return width_map.get(header_name, 25)


def apply_excel_styling(worksheet: Worksheet) -> None:
    """Apply consistent styling to Excel worksheet.

    Applies the following styles:
    - Font: BIZ UDPゴシック (11pt, bold for header)
    - Freeze: First row (header)
    - Background: Alternating blue color for even rows (#E8F4F8)
    - Text wrapping: Enabled for all cells
    - Column width: Auto-adjusted based on content type

    Args:
        worksheet: openpyxl worksheet object to style
    """
    # Define styles
    font = Font(name="BIZ UDPゴシック", size=11)
    header_font = Font(name="BIZ UDPゴシック", size=11, bold=True)
    even_row_fill = PatternFill(
        start_color="E8F4F8", end_color="E8F4F8", fill_type="solid"
    )
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    # Freeze header row (first row)
    worksheet.freeze_panes = "A2"

    # Get header names for column width estimation
    header_names = {}
    if worksheet.max_row > 0:
        for cell in worksheet[1]:
            if cell.value:
                header_names[cell.column] = str(cell.value).strip()

    # Apply styles to all cells
    for row_idx, row in enumerate(worksheet.iter_rows(), start=1):
        is_header = row_idx == 1
        is_even = row_idx % 2 == 0

        for cell in row:
            # Apply font
            cell.font = header_font if is_header else font

            # Apply text wrapping to all cells
            cell.alignment = wrap_alignment

            # Apply background color to even rows (excluding header)
            if not is_header and is_even:
                cell.fill = even_row_fill

    # Set column widths based on header names
    for col_idx, header_name in header_names.items():
        col_letter = get_column_letter(col_idx)
        estimated_width = _estimate_column_width(header_name)
        worksheet.column_dimensions[col_letter].width = estimated_width


__all__ = ["apply_excel_styling"]
