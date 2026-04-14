from __future__ import annotations

import io
import warnings
from collections.abc import Iterable

import pandas as pd

from src.common.config import BopConfig
from src.common.text_utils import sanitize

from .constants import METADATA_COLUMN_LABELS
from .models import BopDataset, PartColumn
from .settings import get_bop_config

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as exc:  # pragma: no cover - dependency guard
    raise ImportError(
        "openpyxl が見つかりません。`pip install -r requirements.txt` を実行して "
        "必須依存関係をインストールしてください。"
    ) from exc

# Deduplicate noisy warning: openpyxl drops unsupported WMF images on load.
warnings.filterwarnings(
    "once",
    category=UserWarning,
    message="wmf image format is not supported.*",
    module="openpyxl.reader.drawings",
)


def _find_row_index(df: pd.DataFrame, keywords: Iterable[str]) -> int:
    for idx, row in df.iterrows():
        values = [sanitize(v) for v in row.tolist() if not pd.isna(v)]
        if all(any(keyword in cell for cell in values) for keyword in keywords):
            return int(idx)
    raise ValueError(f"指定キーワード{keywords}を含む行が見つかりません")


def _build_block_station_maps(
    raw: pd.DataFrame,
    block_row_idx: int,
    station_row_idx: int,
) -> tuple[dict[int, str], dict[int, str]]:
    block_cache: dict[int, str] = {}
    station_cache: dict[int, str] = {}
    current_block = ""
    current_station = ""
    for col in range(raw.shape[1]):
        block_val = sanitize(raw.iat[block_row_idx, col])
        if block_val:
            current_block = block_val
        block_cache[col] = current_block

        station_val = sanitize(raw.iat[station_row_idx, col])
        if station_val:
            current_station = station_val
        station_cache[col] = current_station
    return block_cache, station_cache


def _collect_part_columns(
    raw: pd.DataFrame,
    metadata_row_idx: int,
    block_cache: dict[int, str],
    station_cache: dict[int, str],
    skip_patterns: tuple[str, ...],
) -> tuple[list[PartColumn], int | None]:
    part_columns: list[PartColumn] = []
    first_part_col: int | None = None
    for col in range(raw.shape[1]):
        part_label = sanitize(raw.iat[metadata_row_idx, col])
        if not part_label:
            continue
        if part_label.startswith("（") and part_label.endswith("）"):
            continue
        if part_label in METADATA_COLUMN_LABELS:
            continue
        if any(pattern in part_label for pattern in skip_patterns):
            continue
        block = block_cache.get(col, "")
        station = station_cache.get(col, "")
        if not block or not station:
            continue
        key = f"{block}::{station}::col{col:03d}"
        part_columns.append(
            PartColumn(
                key=key,
                col_idx=col,
                block=block,
                station=station,
                part_label=part_label,
            )
        )
        if first_part_col is None or col < first_part_col:
            first_part_col = col
    return part_columns, first_part_col


def _extract_metadata_headers(
    raw: pd.DataFrame, metadata_row_idx: int, first_part_col: int
) -> list[tuple[int, str]]:
    headers: list[tuple[int, str]] = []
    for col in range(first_part_col):
        header_value = sanitize(raw.iat[metadata_row_idx, col])
        if header_value:
            headers.append((col, header_value.replace(" ", "")))
    return headers


def _build_variant_label(metadata: dict[str, str], offset: int) -> str:
    key_fields = [
        metadata.get("流動ライン"),
        metadata.get("車型"),
        metadata.get("車種"),
        metadata.get("ユニット"),
        metadata.get("HVAC"),
    ]
    key_text = "-".join(filter(None, key_fields))
    if key_text:
        return key_text
    return f"variant-{offset:03d}"


def _build_variant_records(
    raw: pd.DataFrame,
    data_start_idx: int,
    metadata_headers: list[tuple[int, str]],
    part_columns: list[PartColumn],
) -> tuple[list[dict[str, str]], list[dict[str, str]], list[str], dict[str, str]]:
    metadata_records: list[dict[str, str]] = []
    part_records: list[dict[str, str]] = []
    variant_labels: list[str] = []
    annotations: dict[str, str] = {}
    first_metadata_header = metadata_headers[0][1] if metadata_headers else ""

    def _resolve_vertical_reference(column_key: str) -> str:
        for previous_records in reversed(part_records):
            cached_value = previous_records.get(column_key, "")
            if cached_value != "↑":
                return cached_value
        return ""

    for row_idx in range(data_start_idx, raw.shape[0]):
        row_series = raw.iloc[row_idx]
        if all(pd.isna(val) for val in row_series.values):
            continue

        metadata_dict: dict[str, str] = {}
        for col, header in metadata_headers:
            metadata_dict[header] = sanitize(row_series[col])

        if (
            first_metadata_header
            and metadata_dict.get(first_metadata_header) == "形状の特長"
        ):
            for part_col in part_columns:
                value = sanitize(row_series[part_col.col_idx])
                if value:
                    annotations[part_col.key] = value
            continue

        variant_label = _build_variant_label(metadata_dict, len(metadata_records))
        variant_labels.append(variant_label)
        metadata_records.append(metadata_dict)

        parts_dict: dict[str, str] = {}
        for part_col in part_columns:
            value = sanitize(row_series[part_col.col_idx])
            if value == "↑":
                value = _resolve_vertical_reference(part_col.key)
            parts_dict[part_col.key] = value
        part_records.append(parts_dict)
    return metadata_records, part_records, variant_labels, annotations


def _parse_bop_raw(
    raw: pd.DataFrame,
    skip_patterns: tuple[str, ...],
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, str]]:
    block_row_idx = _find_row_index(raw, ["BL"])
    station_row_idx = _find_row_index(raw, ["ST"])
    metadata_row_idx = _find_row_index(raw, ["流動", "ユニット"])
    data_start_idx = metadata_row_idx + 1

    block_cache, station_cache = _build_block_station_maps(
        raw, block_row_idx, station_row_idx
    )
    part_columns, first_part_col = _collect_part_columns(
        raw,
        metadata_row_idx,
        block_cache,
        station_cache,
        skip_patterns,
    )

    if first_part_col is None:
        raise ValueError(
            "部品カラムが認識できませんでした。ヘッダ構造を確認してください。"
        )

    metadata_headers = _extract_metadata_headers(raw, metadata_row_idx, first_part_col)
    metadata_records, part_records, variant_labels, annotations = (
        _build_variant_records(
            raw,
            data_start_idx,
            metadata_headers,
            part_columns,
        )
    )

    metadata_df = pd.DataFrame(metadata_records, index=variant_labels)
    parts_df = pd.DataFrame(part_records, index=variant_labels)
    column_catalog = pd.DataFrame(
        [
            {
                "column_key": part.key,
                "block": part.block,
                "station": part.station,
                "part_label": part.part_label,
                "column_index": part.col_idx,
            }
            for part in part_columns
        ]
    ).set_index("column_key")
    annotations = {str(key): value for key, value in annotations.items()}
    return metadata_df, parts_df, column_catalog, annotations


def _propagate_vertical_merges(ws: Worksheet, raw: pd.DataFrame) -> pd.DataFrame:
    updated = raw.copy()
    for cell_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = cell_range.bounds
        if max_row <= min_row:
            continue
        base_value = ws.cell(row=min_row, column=min_col).value
        if base_value is None:
            continue
        for row in range(min_row + 1, max_row + 1):
            for col in range(min_col, max_col + 1):
                updated.iat[row - 1, col - 1] = base_value
    return updated


def load_bop_master(data: bytes, config: BopConfig | None = None) -> BopDataset:
    conf = config or get_bop_config()
    raw = pd.read_excel(io.BytesIO(data), header=None)
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="wmf image format is not supported so the image is being dropped",
            category=UserWarning,
        )
        wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    raw = _propagate_vertical_merges(ws, raw)
    metadata_df, parts_df, column_catalog, annotations = _parse_bop_raw(
        raw, conf.part_skip_patterns
    )
    return BopDataset(
        metadata=metadata_df,
        parts=parts_df,
        column_catalog=column_catalog,
        annotations=annotations,
    )


__all__ = [
    "_build_block_station_maps",
    "_collect_part_columns",
    "_extract_metadata_headers",
    "_build_variant_records",
    "load_bop_master",
]
